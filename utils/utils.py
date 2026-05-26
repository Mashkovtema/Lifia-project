from openpyxl.styles import PatternFill, Alignment, Font
import openpyxl
import datetime
import logging


async def process_time_settings(current_hour: int, current_minute: int, action: str, flag: str):
    """Настройка времени"""
    logging.info('process_time_settings')

    if action == 'up':
        if flag == 'hour':
            if current_hour == 23:
                current_hour = 0
            else:
                current_hour += 1
        if flag == 'minute':
            if current_minute == 55:
                current_hour += 1
                if current_hour > 23:
                    current_hour = 0
                current_minute = 0
            else:
                current_minute += 5
    else:
        if flag == 'hour':
            if current_hour == 0:
                current_hour = 23
            else:
                current_hour -= 1
        if flag == 'minute':
            if current_minute == 0:
                current_hour -= 1
                if current_hour < 0:
                    current_hour = 23
                current_minute = 55
            else:
                current_minute -= 5

    # Форматируем в строки с ведущим нулём
    return f"{current_hour:02d}", f"{current_minute:02d}"


async def add_sticker_to_category(category: str) -> str:
    """Добавление эмодзи к напоминанию"""
    logging.info('add_sticker_to_category')
    if category == 'Прием врача':
        return '🏥 Прием врача'

    if category == 'Анализы':
        return '📋 Анализы'

    if category == 'Купить что-то малышу':
        return '🛒 Купить что-то малышу'

    if category == 'Заказать подгузники':
        return '📦 Заказать подгузники'

    if category == 'Пить воду':
        return '💧 Пить воду'

    if category == 'Витамины':
        return '💊 Витамины'

    if category == 'Поесть':
        return '🍲 Поесть'

    if category == 'Прогулка':
        return '🚶 Прогулка'

    if category == 'Отдохнуть':
        return '😴 Отдохнуть'


async def validate_int_data(data) -> bool:
    """Проверка на численные данные"""
    logging.info('validate_int_data')
    try:
        cost = int(data)
        if cost > 0:
            return True
        else:
            return False
    except:
        return False


async def create_users_table(users_data: list) -> None:
    """Создание ексель файла с данными пользователей"""
    logging.info('create_users_table')
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    center_aligned_text = Alignment(horizontal="center", vertical="center")

    for column in ['A', 'B', 'C', 'D','E','F','G','H','I', 'J']:
        sheet.column_dimensions[column].width = 15

    # Заголовок
    sheet.append(['№', 'user_id', 'username', 'Имя', 'Неделя', 'Кол-во бонусов', 'Статус', 'Тип подписки', 'Конец подписки', 'Часовая зона'])
    sheet.append([])

    current_row_number = sheet.max_row

    # Выравнивание по центру
    for column in range(1, 15):
        cell = sheet.cell(row=current_row_number, column=column)
        cell.alignment = center_aligned_text

    # Заполнение данных
    for user in users_data:
        user = user.__dict__
        sheet.append([
            user['id'],
            user['user_id'],
            user['username'],
            user['name'],
            user['week'],
            user['bonus_cnt'],
            "Уже мама" if user['mom_or_not'] else "Еще не мама",
            "Про" if user['subscription_type'] == 'pro' else "Стандарт",
            user['subscription_date_end'],
            user['time_zone'] if user['time_zone'] >= 0 else f'-{user["time_zone"]}'
        ])

        current_row_number = sheet.max_row

        # Выравнивание по центру
        for column in range(1, 15):
            cell = sheet.cell(row=current_row_number, column=column)
            cell.alignment = center_aligned_text

    workbook.save('Пользователи.xlsx')


async def calculate_week_ceil(year: int, month: int, day: int) -> int:
    """Подсчет недель с округлением вверх (включая текущую неполную неделю)"""
    logging.info(f'calculate_week_ceil: date={year}-{month}-{day}')

    target_date = datetime.date(year, month, day)
    today = datetime.date.today()

    if target_date > today:
        logging.warning(f'Target date {target_date} is in the future')
        return -1

    # Calculate days difference
    days_diff = (today - target_date).days

    # Calculate weeks with ceiling (round up)
    weeks = (days_diff + 6) // 7  # Integer division with ceiling

    logging.info(f'Days difference: {days_diff}, Weeks (ceil): {weeks}')
    return weeks


async def get_text_by_type(mom_or_not: bool) -> str:
    """Получение текста в зависимости от положения"""
    logging.info('get_text_by_type')
    if mom_or_not:
        text = ('💙 Рождение малыша — это начало совершенно новой жизни\n\n'
                'Вместе с счастьем часто приходят усталость, тревожность и множество новых вопросов.\n\n'
                'Иногда может быть тяжело.\n'
                'Иногда — страшно.\n\n'
                'А иногда просто хочется услышать:\n\n'
                '«Я справляюсь нормально?»\n'
                '«Почему я так устаю?»\n'
                '«Что сейчас происходит с моим организмом?»\n'
                '«Почему малыш плачет?»\n\n'
                'ILIFIA создана, чтобы поддерживать тебя и после родов 🤍\n\n'
                'Здесь ты сможешь:\n\n'
                '👶 лучше понимать малыша и его развитие в первые месяцы жизни\n\n'
                '🤍 спокойнее восстанавливаться после родов и понимать изменения своего организма\n\n'
                '😴 получать советы о сне малыша, режиме и самочувствии\n\n'
                '🍼 узнавать больше о кормлении и уходе за малышом без давления и осуждения\n\n'
                '🌿 получать поддержку в моменты тревоги, усталости и эмоционального перегруза\n\n'
                '🧘 мягко восстанавливать тело и эмоциональное состояние\n\n'
                '📖 сохранять важные моменты первых месяцев материнства\n\n'
                'И самое главное — помнить, что ты не одна 💙\n\n'
                'Я рядом с тобой 24/7')
    else:
        text = ('💙 Беременность — это особенный период жизни\n\n'
                'Вместе с радостью часто приходят тревоги, новые ощущения и сотни вопросов.\n'
                'Иногда может быть страшно. Иногда — одиноко. А иногда просто хочется понять:\n'
                '«Это нормально?»\n'
                '«Почему я так себя чувствую?»\n'
                '«Что сейчас происходит с малышом?»\n\n'
                'ILIFIA создана, чтобы быть рядом с тобой в этот важный период 🤍\n\n'
                'Здесь ты сможешь:\n\n'
                '🌸 получать поддержку и ответы на вопросы о беременности\n\n'
                '👶 узнавать, что происходит с малышом на твоем сроке\n\n'
                '🤍 лучше понимать свое состояние и изменения в организме\n\n'
                '🧘 готовиться к родам спокойнее и увереннее\n\n'
                '📖 сохранять важные моменты беременности в дневнике\n\n'
                'И самое главное — помнить, что ты не одна 💙\n\n'
                'Я рядом с тобой 24/7')

    return text


























