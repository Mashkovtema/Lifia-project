from database.models import async_session, AdminReminds, AiToken, Tarifs, Reviews, Challenges, Users, WeeksTextsBer, WeeksTextsMom
from sqlalchemy import select, or_, and_, delete, func, case, cast, Integer, String
from passlib.context import CryptContext
import logging
import openpyxl

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


async def add_new_admin_remind(category: str, time_type: str, times: str, days: str, comment: str, users_type: str) -> None:
    """Добавление нового напоминания админом"""
    logging.info('add_new_admin_remind')
    async with async_session() as session:
        new_remind = AdminReminds(
            category=category,
            time_type=time_type,
            times=times,
            days=days,
            users_type=users_type,
            comment=comment
        )
        session.add(new_remind)
        await session.commit()


async def get_reminds_by_type(users_type: str) -> list:
    """Получение списка напоминаний по типу пользователей"""
    logging.info('get_reminds_by_type')
    async with async_session() as session:
        data = await session.scalars(select(AdminReminds).where(AdminReminds.users_type == users_type))
        if data:
            return data.all()
        else:
            return []


async def get_remind_by_index(remind_index: int) -> dict:
    """Получение напоминания по индексу"""
    logging.info('get_remind_by_index')
    async with async_session() as session:
        remind = await session.scalar(select(AdminReminds).where(AdminReminds.id == remind_index))
        return remind.__dict__


async def delete_remind_by_index(remind_index: int) -> None:
    """Удаление напоминания по индексу"""
    logging.info('delete_remind_by_index')
    async with async_session() as session:
        remind = await session.scalar(select(AdminReminds).where(AdminReminds.id == remind_index))
        await session.delete(remind)
        await session.commit()


async def update_ai_token(ai_token: str) -> None:
    """Обновление токена к моделям"""
    logging.info('update_ai_token')
    async with async_session() as session:
        token = await session.scalar(select(AiToken).where(AiToken.id == 1))
        if token:
            token.token = ai_token
            await session.commit()
        else:
            new_token = AiToken(
                token=ai_token
            )
            session.add(new_token)
            await session.commit()


async def get_token() -> str:
    """Получение токена"""
    logging.info('get_token')
    async with async_session() as session:
        token = await session.scalar(select(AiToken.token).where(AiToken.id == 1))
        if token:
            return token
        else:
            return 'Не установлен'


async def get_tarif_type(tarif_type: str):
    """Получение тарифа по типу"""
    logging.info('get_tarif_type')
    async with async_session() as session:
        tarif_data = await session.scalar(select(Tarifs).where(Tarifs.tarif_type == tarif_type))
        if tarif_data:
            return tarif_data.__dict__
        else:
            return None


async def add_new_tarif(tarif_type: str, name: str, message_cnt: int, photo: str, cost: int) -> None:
    """Добавление нового тарифа"""
    logging.info('add_new_tarif')
    async with async_session() as session:
        new_tarif = Tarifs(
            tarif_type=tarif_type,
            cost=cost,
            name=name,
            message_cnt=message_cnt,
            photo=photo
        )
        session.add(new_tarif)
        await session.commit()


async def update_tarif_data(tarif_type: str, parametr: str, value: str) -> None:
    """Обновление данных о тарифах"""
    logging.info('update_tarif_data')
    async with async_session() as session:
        tarif_data = await session.scalar(select(Tarifs).where(Tarifs.tarif_type == tarif_type))
        if parametr == 'name':
            tarif_data.name = value
        if parametr == 'cost':
            tarif_data.cost = value
        if parametr == 'message-cnt':
            tarif_data.message_cnt = value
        if parametr == 'photo':
            tarif_data.photo = value
        await session.commit()


async def get_reviews_for_moderation() -> list:
    """Получение отзывов для модерации"""
    logging.info('get_reviews_for_moderation')
    async with async_session() as session:
        reviews = await session.scalars(select(Reviews).where(Reviews.moderation == False))
        if reviews:
            return reviews.all()
        else:
            return []


async def confirm_review(index: int) -> None:
    """Подтверждение отзыва"""
    logging.info('confirm_review')
    async with async_session() as session:
        review = await session.scalar(select(Reviews).where(Reviews.id == index))
        review.moderation = True
        await session.commit()


async def delete_review(index: int) -> None:
    """Удаление отзыва"""
    logging.info('delete_review')
    async with async_session() as session:
        review = await session.scalar(select(Reviews).where(Reviews.id == index))
        await session.delete(review)
        await session.commit()


async def get_chelenges_by_category(category: str, week: int) -> list:
    """Получение списка задач по категории"""
    logging.info('get_chelenges_by_category')
    async with async_session() as session:
        challenges_data = await session.scalars(select(Challenges).where(Challenges.category == category, Challenges.week == week))
        if challenges_data:
            return challenges_data.all()
        else:
            return []


async def add_new_challenge(challenge_data: dict) -> None:
    """Добавление нового задания"""
    logging.info('add_new_challenge')
    async with async_session() as session:
        new_challenge = Challenges(
            category=challenge_data['category'],
            name=challenge_data['name'],
            bonus_cnt=challenge_data['bonus_cnt'],
            week=challenge_data['week']
        )
        session.add(new_challenge)
        await session.commit()


async def get_challenge_by_index(index: int) -> dict:
    """Получение задачи по индексу"""
    logging.info('get_challenge_by_index')
    async with async_session() as session:
        challenge = await session.scalar(select(Challenges).where(Challenges.id == index))
        return challenge.__dict__


async def delete_challenge(index: str) -> None:
    """Удаление Задачи"""
    logging.info('delete_challenge')
    async with async_session() as session:
        challenge = await session.scalar(select(Challenges).where(Challenges.id == index))
        await session.delete(challenge)
        await session.commit()


async def get_user_ids_for_newsletter(type: str) -> list:
    """Получение спсика пользователей для рассылки"""
    logging.info('get_user_ids_for_newsletter')
    async with async_session() as session:
        if type == 'not-mom':
            users = await session.scalars(select(Users.user_id).where(Users.mom_or_not == False))
        elif type == 'mom':
            users = await session.scalars(select(Users.user_id).where(Users.mom_or_not == True))
        else:
            users = await session.scalars(select(Users.user_id))

        if users:
            return users.all()
        else:
            return []


async def get_all_users_data() -> list:
    """Получение нформации о пользователях"""
    logging.info('get_all_users_data')
    async with async_session() as session:
        users_data = await session.scalars(select(Users))
        if users_data:
            return users_data.all()
        else:
            return []


async def update_weeks_text(file_in_memory, type: str) -> None:
    """Обновление текстов по дням беременности"""
    logging.info('update_weeks_text')
    async with async_session() as session:
        if type == 'Беременные':
            await session.execute(delete(WeeksTextsBer))
            await session.commit()

            workbook = openpyxl.load_workbook(file_in_memory)

            sheet = workbook.active

            for i in range(2, sheet.max_row + 1):
                day_data = sheet[f'A{i}'].value
                text_data = sheet[f'D{i}'].value

                new_week_text_data = WeeksTextsBer(
                    day=day_data,
                    text=text_data
                )
                session.add(new_week_text_data)

            await session.commit()

        else:
            await session.execute(delete(WeeksTextsMom))
            await session.commit()

            workbook = openpyxl.load_workbook(file_in_memory)

            sheet = workbook.active

            for i in range(2, sheet.max_row + 1):
                day_data = sheet[f'A{i}'].value
                text_data = sheet[f'D{i}'].value

                new_week_text_data = WeeksTextsMom(
                    day=day_data,
                    text=text_data
                )
                session.add(new_week_text_data)

            await session.commit()












